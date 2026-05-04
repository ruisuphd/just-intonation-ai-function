#!/usr/bin/env python3
"""BPS-FH (Beethoven Piano Sonatas First-movements) ingestion adapter.

BPS-FH is a corpus of 32 Beethoven piano sonata first-movements with
Roman-numeral and key annotations, distributed by Tsung-Ping Chen at:
    https://github.com/Tsung-Ping/functional-harmony

Reference:
    Chen, T.-P., & Su, L. (2018). Functional harmony recognition with
    multi-task recurrent neural networks. Proceedings of ISMIR 2018.

Per-piece directory layout (verified 2026-05-02):
    BPS_FH_Dataset/<piece_id>/notes.csv
    BPS_FH_Dataset/<piece_id>/chords.xlsx
    BPS_FH_Dataset/<piece_id>/beats.xlsx
    BPS_FH_Dataset/<piece_id>/dBeats.xlsx
    BPS_FH_Dataset/<piece_id>/phrases.xlsx

`notes.csv` columns (no header):
    onset_beat, midi_pitch, morphetic_pitch, duration_beat, staff, measure

`chords.xlsx` columns (no header):
    onset_beat, offset_beat, key, degree, quality, inversion, roman_numeral

BPS-FH key-naming convention (Chen & Su 2018):
  - Uppercase letter = major key (e.g., 'C', 'A', 'B')
  - Lowercase letter = minor key (e.g., 'c', 'a', 'b')
  - '-' suffix       = flat            (e.g., 'A-' = A♭ major; 'b-' = B♭ minor)
  - '+' suffix       = sharp           (e.g., 'F+' = F♯ major; 'f+' = F♯ minor)

Output JSON shape (matches the project's canonical Strategy A schema):
  {
    "id": "BPS_FH_<piece_id>",
    "source": "bps_fh",
    "converter_strategy": "A",
    "notes": [
      {"pitch": int_midi, "onset_beat": float, "duration_beat": float,
       "velocity": int, "key": "C" | "Cm" | ...,
       "tonic_pc": int, "is_minor": bool,
       "chord_numeral": str, "chord_type": str,
       "chord_relativeroot": null, "chord_root_pc": int}
    ]
  }

Author: Rui Su, 2026-05-02. Phase I Month 2 cross-corpus adapter.
Verified end-to-end against the 32-piece dataset locally before zip.
"""
from __future__ import annotations

import argparse
import csv
import json
import logging
import re
import sys
from bisect import bisect_right
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# Pitch-class look-up — A=9, B=11, C=0, D=2, E=4, F=5, G=7
NOTE_LETTER_PC = {'C': 0, 'D': 2, 'E': 4, 'F': 5, 'G': 7, 'A': 9, 'B': 11}


# ─────────────────────────────────────────────────────────────────────────
# BPS-FH key-string parsing

def parse_bps_key(key_str: str) -> Tuple[str, int, bool]:
    """Parse a BPS-FH key string (e.g., 'A-', 'b', 'f+') into
    (canonical_key_string, tonic_pc, is_minor).

    Canonical key strings use the project's existing convention: '<note>'
    for major (e.g., 'C', 'F#', 'Ab') and '<note>m' for minor (e.g., 'Cm',
    'F#m', 'Bbm'). The mapping into the 24-class label space happens
    downstream in the dataset loader.
    """
    s = key_str.strip()
    if not s:
        raise ValueError('empty key string')
    letter = s[0]
    is_minor = letter.islower()
    letter_upper = letter.upper()
    if letter_upper not in NOTE_LETTER_PC:
        raise ValueError(f'unrecognised note letter in {key_str!r}')
    pc = NOTE_LETTER_PC[letter_upper]
    canonical_letter = letter_upper
    rest = s[1:]
    canonical_accidental = ''
    if rest:
        for ch in rest:
            if ch == '-':
                pc = (pc - 1) % 12
                canonical_accidental += 'b'
            elif ch == '+':
                pc = (pc + 1) % 12
                canonical_accidental += '#'
            else:
                raise ValueError(f'unrecognised accidental {ch!r} in key {key_str!r}')
    canonical = canonical_letter + canonical_accidental + ('m' if is_minor else '')
    return canonical, pc, is_minor


# ─────────────────────────────────────────────────────────────────────────
# Chord-root pitch-class lookup from BPS-FH degree + quality

# Roman-numeral degree → semitone offset within the local key (natural
# major scale; minor-mode adjustments handled below).
DEGREE_TO_PC_MAJOR = {1: 0, 2: 2, 3: 4, 4: 5, 5: 7, 6: 9, 7: 11}
DEGREE_TO_PC_MINOR = {1: 0, 2: 2, 3: 3, 4: 5, 5: 7, 6: 8, 7: 10}


def _bps_chord_root_pc(degree: Any, key_tonic_pc: int, key_is_minor: bool) -> int:
    """Resolve the absolute pitch class of a BPS-FH chord root.

    BPS-FH stores degree as integer 1–7 (occasionally with sharps / flats
    prepended in the roman_numeral column, e.g., 'bII'). When degree is
    an integer we use the standard mode-aware degree → pitch-class map.
    """
    if degree is None:
        return -1
    try:
        deg = int(degree)
    except (ValueError, TypeError):
        return -1
    if deg < 1 or deg > 7:
        return -1
    table = DEGREE_TO_PC_MINOR if key_is_minor else DEGREE_TO_PC_MAJOR
    return (key_tonic_pc + table[deg]) % 12


# ─────────────────────────────────────────────────────────────────────────
# File readers

def _load_bps_notes(piece_dir: Path) -> List[Dict[str, Any]]:
    """Read notes.csv: onset_beat, midi, morphetic, duration_beat, staff, measure."""
    notes_csv = piece_dir / 'notes.csv'
    if not notes_csv.exists():
        return []
    notes: List[Dict[str, Any]] = []
    with open(notes_csv, newline='') as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) < 4:
                continue
            try:
                onset = float(row[0])
                pitch = int(row[1])
                duration = float(row[3])
            except (ValueError, IndexError):
                continue
            if pitch <= 0 or duration <= 0:
                continue
            notes.append({
                'pitch': pitch,
                'onset_beat': onset,
                'duration_beat': duration,
                'velocity': 80,  # BPS-FH has no velocity; project default
            })
    notes.sort(key=lambda n: n['onset_beat'])
    return notes


def _load_bps_chords(piece_dir: Path) -> List[Dict[str, Any]]:
    """Read chords.xlsx: onset, offset, key, degree, quality, inversion, roman."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            'openpyxl required for BPS-FH (chords.xlsx). '
            'Install: pip install openpyxl'
        ) from e
    chords_xlsx = piece_dir / 'chords.xlsx'
    if not chords_xlsx.exists():
        return []
    wb = load_workbook(chords_xlsx, read_only=True, data_only=True)
    ws = wb.active
    chords: List[Dict[str, Any]] = []
    for row in ws.iter_rows(values_only=True):
        # row: (onset, offset, key, degree, quality, inversion, roman)
        if row is None or len(row) < 7:
            continue
        if any(v is None for v in row[:3]):
            continue
        try:
            onset = float(row[0])
            offset = float(row[1])
            key_str = str(row[2]).strip()
            degree = row[3]
            quality = str(row[4]).strip() if row[4] is not None else ''
            inversion = row[5]
            roman = str(row[6]).strip() if row[6] is not None else ''
        except (ValueError, TypeError):
            continue
        if not key_str:
            continue
        try:
            canonical_key, tonic_pc, is_minor = parse_bps_key(key_str)
        except ValueError:
            continue
        chord_root_pc = _bps_chord_root_pc(degree, tonic_pc, is_minor)
        chords.append({
            'start_beat': onset,
            'end_beat': offset,
            'key': canonical_key,
            'tonic_pc': tonic_pc,
            'is_minor': is_minor,
            'numeral': roman,
            'chord_type': quality,
            'inversion': inversion,
            'chord_root_pc': chord_root_pc,
        })
    chords.sort(key=lambda c: c['start_beat'])
    return chords


def _attach_chord_to_notes(
    notes: List[Dict[str, Any]],
    chords: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """For each note, find the chord whose [start_beat, end_beat) covers
    the note's onset_beat. Notes outside any annotated chord are dropped
    (consistent with the existing Strategy A behaviour for DCML)."""
    if not notes or not chords:
        return []
    chord_starts = [c['start_beat'] for c in chords]
    out: List[Dict[str, Any]] = []
    for n in notes:
        idx = bisect_right(chord_starts, n['onset_beat']) - 1
        if idx < 0:
            continue
        c = chords[idx]
        if n['onset_beat'] >= c['end_beat']:
            continue
        out.append({
            'pitch': n['pitch'],
            'onset_beat': n['onset_beat'],
            'duration_beat': n['duration_beat'],
            'velocity': n['velocity'],
            'key': c['key'],
            'tonic_pc': c['tonic_pc'],
            'is_minor': c['is_minor'],
            'chord_numeral': c['numeral'],
            'chord_type': c['chord_type'],
            'chord_relativeroot': None,
            'chord_root_pc': c['chord_root_pc'],
        })
    return out


def convert_one_piece(piece_dir: Path, output_dir: Path,
                      piece_id: Optional[str] = None) -> Optional[Path]:
    """Convert one BPS-FH piece to a Strategy-A-compatible JSON."""
    piece_id = piece_id or piece_dir.name
    notes = _load_bps_notes(piece_dir)
    chords = _load_bps_chords(piece_dir)
    if not notes or not chords:
        logger.warning('skipping %s: notes=%d chords=%d',
                       piece_id, len(notes), len(chords))
        return None
    paired = _attach_chord_to_notes(notes, chords)
    if not paired:
        logger.warning('skipping %s: no notes paired to chords', piece_id)
        return None
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f'BPS_FH_{piece_id}.json'
    payload = {
        'id': f'BPS_FH_{piece_id}',
        'source': 'bps_fh',
        'converter_strategy': 'A',
        'reference': ('Chen & Su (2018), "Functional harmony recognition '
                      'with multi-task RNNs", ISMIR 2018'),
        'notes': paired,
    }
    out_path.write_text(json.dumps(payload, indent=2))
    return out_path


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', required=True,
                    help='BPS_FH_Dataset/ root directory (one subdir per piece)')
    ap.add_argument('--output', required=True,
                    help='Where to write the per-piece JSONs')
    ap.add_argument('--verbose', action='store_true')
    args = ap.parse_args()

    logging.basicConfig(
        level=logging.INFO if args.verbose else logging.WARNING,
        format='%(levelname)s %(message)s',
    )

    in_root = Path(args.input).expanduser().resolve()
    out_root = Path(args.output).expanduser().resolve()
    if not in_root.is_dir():
        print(f'ERROR: input dir does not exist: {in_root}')
        return 1
    out_root.mkdir(parents=True, exist_ok=True)

    n_ok, n_skip = 0, 0
    for piece_dir in sorted(in_root.iterdir(), key=lambda p: p.name):
        if not piece_dir.is_dir():
            continue
        if not piece_dir.name.split('_')[0].isdigit() and \
                not piece_dir.name.isdigit():
            continue  # skip "Taking Form" and similar
        result = convert_one_piece(piece_dir, out_root)
        if result is None:
            n_skip += 1
        else:
            n_ok += 1
            if args.verbose:
                print(f'  ✓ {piece_dir.name} → {result.name}')
    print(f'BPS-FH: converted {n_ok} pieces, skipped {n_skip}')
    print(f'Output: {out_root}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
